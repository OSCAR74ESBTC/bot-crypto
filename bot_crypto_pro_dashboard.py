import asyncio
import requests
import pandas as pd
import numpy as np
import os

TOKEN = os.getenv("TOKEN")
CHAT_ID = os.getenv("CHAT_ID")
from ta.momentum import RSIIndicator
from datetime import datetime
from telegram import Bot
import schedule
import time
import matplotlib.pyplot as plt
import io
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image

# =========================
# CONFIGURACIÓN
# =========================
TOKEN = "8608156480:AAE9Ga9eCPIlIw6mRoopvvRANxhlsPsDJqs"       # Tu token Telegram
CHAT_ID = "1084796259"   # Tu chat ID
bot = Bot(token=TOKEN)

cryptos = {
    "BTC": "bitcoin",
    "ETH": "ethereum",
    "SOL": "solana",
    "XRP": "ripple"
}

timeframes = {
    "5m": "5min",
    "15m": "15min",
    "30m": "30min",
    "1h": "60min",
    "1d": "1D"
}

take_profit_pct = 0.03
stop_loss_pct = 0.02
capital_inicial = 5.0
excel_file = "trading_dashboard.xlsx"

# =========================
# FUNCIONES DE DATOS
# =========================
def get_data(coin_id):
    url = f"https://api.coingecko.com/api/v3/coins/{coin_id}/market_chart"
    params = {"vs_currency": "eur", "days": "1"}
    try:
        response = requests.get(url, params=params)
        data = response.json()
        if "prices" not in data:
            print(f"❌ Error API con {coin_id}: {data}")
            return None
        df = pd.DataFrame(data["prices"], columns=["timestamp", "price"])
        df["timestamp"] = pd.to_datetime(df["timestamp"], unit='ms')
        df.set_index("timestamp", inplace=True)
        return df
    except Exception as e:
        print(f"❌ Error conexión: {e}")
        return None

def connors_rsi(df):
    rsi = RSIIndicator(df["price"], window=3).rsi()
    df["change"] = df["price"].diff()
    df["updown"] = np.where(df["change"] > 0, 1, -1)
    streak = []
    count = 0
    for val in df["updown"]:
        count = count + 1 if val == 1 else count - 1
        streak.append(count)
    streak = pd.Series(streak)
    rsi_streak = RSIIndicator(streak, window=2).rsi()
    return (rsi + rsi_streak) / 2

def analyze(df):
    rsi = RSIIndicator(df["price"], window=14).rsi()
    crsi = connors_rsi(df)
    last_rsi = rsi.iloc[-1]
    last_crsi = crsi.iloc[-1]
    price = df["price"].iloc[-1]
    signal = "HOLD"
    if last_rsi < 30 and last_crsi < 20:
        signal = "BUY"
    elif last_rsi > 70 and last_crsi > 80:
        signal = "SELL"
    return price, last_rsi, last_crsi, signal

# =========================
# FUNCIONES EXCEL
# =========================
def save_to_excel(row):
    try:
        df = pd.read_excel(excel_file)
        df = pd.concat([df, pd.DataFrame([row])])
    except:
        df = pd.DataFrame([row])
    df.to_excel(excel_file, index=False)

def save_plot_to_excel(buf, crypto):
    """Guardar gráfico en la hoja de cada crypto, reemplazando el anterior."""
    try:
        wb = load_workbook(excel_file)
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = crypto
        wb.save(excel_file)
        wb = load_workbook(excel_file)

    if crypto in wb.sheetnames:
        ws = wb[crypto]
        ws._images = []  # borrar imagen anterior
    else:
        ws = wb.create_sheet(crypto)

    img = Image(buf)
    img.width = 600
    img.height = 300
    ws.add_image(img, "A2")  # siempre en A2 para reemplazar anterior
    wb.save(excel_file)

# =========================
# FUNCIONES GRAFICO
# =========================
def generate_plot(df_tf, crypto, timeframe, price, sl, tp, signal):
    plt.figure(figsize=(8,4))
    plt.plot(df_tf.index, df_tf["price"], label="Precio", color="blue")
    if signal == "BUY":
        plt.scatter(df_tf.index[-1], price, color="green", s=100, label="BUY")
    elif signal == "SELL":
        plt.scatter(df_tf.index[-1], price, color="red", s=100, label="SELL")
    plt.axhline(sl, color="orange", linestyle="--", label="Stop Loss")
    plt.axhline(tp, color="purple", linestyle="--", label="Take Profit")
    plt.title(f"{crypto} - {timeframe}")
    plt.ylabel("Precio (€)")
    plt.legend()
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    plt.close()
    buf.seek(0)
    return buf

# =========================
# FUNCION TELEGRAM
# =========================
async def send_telegram(message):
    try:
        await bot.send_message(chat_id=CHAT_ID, text=message)
    except Exception as e:
        print(f"❌ Error Telegram: {e}")

# =========================
# FUNCION PRINCIPAL
# =========================
positions = {}  # Guardar estado de cada crypto y timeframe

async def run_bot():
    for name, coin in cryptos.items():
        df = get_data(coin)
        if df is None:
            continue
        for tf_name, tf_alias in timeframes.items():
            df_tf = df["price"].resample(tf_alias).last().dropna().to_frame()
            if len(df_tf) < 3:
                continue
            price, rsi, crsi, signal = analyze(df_tf)
            sl = price * (1 - stop_loss_pct)
            tp = price * (1 + take_profit_pct)
            key = f"{name}_{tf_name}"

            if key not in positions:
                positions[key] = {"capital": capital_inicial, "price_buy": 0, "active": False}

            pos = positions[key]

            if signal == "BUY" and not pos["active"]:
                pos["price_buy"] = price
                pos["active"] = True
                pnl = None
                price_sell = None
            elif signal == "SELL" and pos["active"]:
                price_sell = price
                pnl = (price_sell - pos["price_buy"]) * (pos["capital"] / pos["price_buy"])
                pos["active"] = False
                pos["price_buy"] = 0
            else:
                pnl = None
                price_sell = None

            # Construir mensaje completo para CMD y Telegram
            msg_op = (f"{name} | {tf_name}\n"
                      f"Precio: {price:.2f}€ | RSI: {rsi:.2f} | Connors RSI: {crsi:.2f}\n"
                      f"Señal: {signal} | SL: {sl:.2f}€ | TP: {tp:.2f}€")
            if pnl is not None:
                msg_op += f"\nPnL: {pnl:.2f}€"

            # Guardar en Excel
            save_to_excel({
                "time": datetime.now(),
                "crypto": name,
                "timeframe": tf_name,
                "price": price,
                "RSI": rsi,
                "Connors_RSI": crsi,
                "signal": signal,
                "SL": sl,
                "TP": tp,
                "Capital": pos["capital"],
                "Precio_Compra": pos["price_buy"] if pos["active"] else None,
                "Precio_Venta": price_sell,
                "PnL": pnl
            })

            # Guardar gráfico actualizado
            buf = generate_plot(df_tf, name, tf_name, price, sl, tp, signal)
            save_plot_to_excel(buf, name)

            # CMD y Telegram
            print(msg_op)
            await send_telegram(msg_op)

# =========================
# EJECUTAR CADA 5 MINUTOS
# =========================
def job():
    asyncio.run(run_bot())

schedule.every(5).minutes.do(job)

print("BOT CRYPTO PRO FULL + CAPITAL + DASHBOARD PROFESIONAL INICIADO…")
while True:
    schedule.run_pending()
    time.sleep(5)
